"""
XLSX Reader Plugin for extracting data from Excel files.
"""

import base64
import io
from dataclasses import dataclass
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class XLSXExtractionResult:
    """Result of XLSX extraction containing data and metadata."""
    
    sheets: Dict[str, List[List[Any]]]  # Sheet name -> List of rows (each row is a list of cell values)
    metadata: Dict[str, Any]
    success: bool
    error_message: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        """Convert result to dictionary format."""
        return {
            "sheets": self.sheets,
            "metadata": self.metadata,
            "success": self.success,
            "error_message": self.error_message,
        }


class XLSXReaderPlugin:
    """Plugin for reading and extracting data from XLSX files."""

    def extract_xlsx_data(
        self,
        xlsx_content: Optional[str] = None,
        file_path: Optional[str] = None,
        max_rows_per_sheet: int = 1000,
    ) -> XLSXExtractionResult:
        """
        Extract data from an XLSX file.

        Args:
            xlsx_content: Base64 encoded XLSX file content
            file_path: Path to the XLSX file
            max_rows_per_sheet: Maximum number of rows to extract per sheet (default: 1000)

        Returns:
            XLSXExtractionResult containing extracted data and metadata
        """
        try:
            # Load the workbook from either base64 content or file path
            if xlsx_content:
                xlsx_bytes = base64.b64decode(xlsx_content)
                workbook = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
            elif file_path:
                workbook = load_workbook(file_path, read_only=True, data_only=True)
            else:
                return XLSXExtractionResult(
                    sheets={},
                    metadata={},
                    success=False,
                    error_message="Either xlsx_content or file_path must be provided",
                )

            # Extract data from all sheets
            sheets_data: Dict[str, List[List[Any]]] = {}
            sheet_metadata: Dict[str, Dict[str, Any]] = {}

            for sheet_name in workbook.sheetnames:
                sheet: Worksheet = workbook[sheet_name]
                
                # Extract rows from the sheet
                rows_data: List[List[Any]] = []
                row_count = 0
                
                for row in sheet.iter_rows(values_only=True):
                    if row_count >= max_rows_per_sheet:
                        break
                    
                    # Convert row tuple to list, handling None values
                    row_list = [cell if cell is not None else "" for cell in row]
                    rows_data.append(row_list)
                    row_count += 1

                sheets_data[sheet_name] = rows_data
                
                # Collect metadata for each sheet
                sheet_metadata[sheet_name] = {
                    "row_count": row_count,
                    "column_count": sheet.max_column or 0,
                    "total_rows": sheet.max_row or 0,
                    "truncated": row_count >= max_rows_per_sheet,
                }

            # Overall metadata
            metadata = {
                "total_sheets": len(workbook.sheetnames),
                "sheet_names": workbook.sheetnames,
                "sheet_metadata": sheet_metadata,
            }

            workbook.close()

            return XLSXExtractionResult(
                sheets=sheets_data,
                metadata=metadata,
                success=True,
            )

        except Exception as e:
            return XLSXExtractionResult(
                sheets={},
                metadata={},
                success=False,
                error_message=f"Error extracting XLSX data: {str(e)}",
            )

    def extract_xlsx_from_base64(
        self,
        base64_content: str,
        max_rows_per_sheet: int = 1000,
    ) -> XLSXExtractionResult:
        """
        Extract data from a base64 encoded XLSX file.

        Args:
            base64_content: Base64 encoded XLSX file content
            max_rows_per_sheet: Maximum number of rows to extract per sheet

        Returns:
            XLSXExtractionResult containing extracted data and metadata
        """
        return self.extract_xlsx_data(
            xlsx_content=base64_content,
            max_rows_per_sheet=max_rows_per_sheet,
        )

    def extract_xlsx_from_file(
        self,
        file_path: str,
        max_rows_per_sheet: int = 1000,
    ) -> XLSXExtractionResult:
        """
        Extract data from an XLSX file.

        Args:
            file_path: Path to the XLSX file
            max_rows_per_sheet: Maximum number of rows to extract per sheet

        Returns:
            XLSXExtractionResult containing extracted data and metadata
        """
        return self.extract_xlsx_data(
            file_path=file_path,
            max_rows_per_sheet=max_rows_per_sheet,
        )

    def get_xlsx_metadata(
        self,
        xlsx_content: Optional[str] = None,
        file_path: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Get metadata from an XLSX file without extracting full content.

        Args:
            xlsx_content: Base64 encoded XLSX file content
            file_path: Path to the XLSX file

        Returns:
            Dictionary containing metadata about the XLSX file
        """
        try:
            # Load the workbook
            if xlsx_content:
                xlsx_bytes = base64.b64decode(xlsx_content)
                workbook = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
            elif file_path:
                workbook = load_workbook(file_path, read_only=True, data_only=True)
            else:
                return {
                    "success": False,
                    "error_message": "Either xlsx_content or file_path must be provided",
                }

            # Collect metadata
            sheet_info = {}
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_info[sheet_name] = {
                    "row_count": sheet.max_row or 0,
                    "column_count": sheet.max_column or 0,
                }

            metadata = {
                "success": True,
                "total_sheets": len(workbook.sheetnames),
                "sheet_names": workbook.sheetnames,
                "sheets": sheet_info,
            }

            workbook.close()
            return metadata

        except Exception as e:
            return {
                "success": False,
                "error_message": f"Error getting XLSX metadata: {str(e)}",
            }
